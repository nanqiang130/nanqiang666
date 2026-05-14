import os
import io
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════
# 页面配置
# ═══════════════════════════════════════════
st.set_page_config(page_title="辅立码课工具箱", page_icon="🛠️", layout="wide")

# 侧边栏导航
page = st.sidebar.radio("📋 选择功能", [
    "📊 周报数据生成",
    "🎯 目标生考试知识点分析",
    "📝 个人考试知识点分析",
    "📈 个人总表生成"
])

st.sidebar.markdown("---")
st.sidebar.markdown("💡 选择功能后上传对应文件即可")


# ═══════════════════════════════════════════
# 通用工具函数
# ═══════════════════════════════════════════

def wb_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════
# 功能1: 周报数据生成
# ═══════════════════════════════════════════

SUBJECT_ORDER = ["数学", "语文", "英语", "物理", "化学", "生物", "地理", "历史", "政治", "日语", "俄语"]
SUBJECT_RANK = {s: i for i, s in enumerate(SUBJECT_ORDER)}

GREEN_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF")

COLOR_RULES_ZB = [
    ("日均习题任务", 0.9, 0.9, True, False),
    ("完成率", 0.85, 0.80, True, True),
    ("阅卷率", 0.80, 0.80, True, True),
    ("覆盖率", 0.70, None, True, True),
    ("得分率", 0.80, None, True, True),
]


def zb_parse_rate(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val) / 100 if float(val) > 1 else float(val)
    s = str(val).strip().replace("%", "")
    try:
        f = float(s)
        return f / 100 if f > 1 else f
    except:
        return None


def zb_get_color_rule(col_name):
    if not col_name:
        return None
    name = col_name.strip()
    if "以上" in name or "以下" in name or "标绿" in name or "标红" in name:
        return None
    for keyword, green_above, red_below, white_font, is_percent in COLOR_RULES_ZB:
        if keyword in name:
            return (green_above, red_below, white_font, is_percent)
    return None


def zb_detect_cols(ws, header_row=3):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            m[v.strip()] = c
    return m


def zb_sort_rows(data, col_names, subj_col="学科", sort_col="习题任务"):
    si = col_names.index(subj_col) if subj_col in col_names else 0
    ki = col_names.index(sort_col) if sort_col in col_names else -1

    def key(row):
        subj = row[si] or ""
        rank = SUBJECT_RANK.get(subj, 99)
        try:
            val = float(row[ki]) if ki >= 0 and row[ki] not in (None, "") else 0
        except:
            val = 0
        return (rank, -val)

    return sorted(data, key=key)


def zb_write_sheet(wb, name, headers, data, auto_color=True):
    ws = wb.create_sheet(name)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    for ri, row in enumerate(data, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    if auto_color:
        col_rules = {}
        for ci, h in enumerate(headers, 1):
            rule = zb_get_color_rule(h)
            if rule:
                col_rules[ci] = rule
        for ri, row in enumerate(data, 2):
            for ci, val in enumerate(row, 1):
                if ci in col_rules:
                    green_above, red_below, use_white, is_percent = col_rules[ci]
                    rate = zb_parse_rate(val) if is_percent else (float(val) if val not in (None, "") else None)
                    cell = ws.cell(ri, ci)
                    if rate is not None:
                        if green_above is not None and rate >= green_above:
                            cell.fill = GREEN_FILL
                            if use_white: cell.font = WHITE_FONT
                        elif red_below is not None and rate < red_below:
                            cell.fill = RED_FILL
                            if use_white: cell.font = WHITE_FONT


def zb_generate(input_bytes):
    wb_in = openpyxl.load_workbook(io.BytesIO(input_bytes))
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    ws_t = wb_in["教师教学详情"]
    t_cols = zb_detect_cols(ws_t)
    teacher_rows = []
    for r in range(4, ws_t.max_row + 1):
        row = {name: ws_t.cell(r, c).value for name, c in t_cols.items()}
        if row.get("姓名"):
            teacher_rows.append(row)

    s1_cols = ["姓名", "学科", "任务量", "习题任务", "日均习题任务", "生均题量", "完成率", "阅卷率"]
    s1_data = zb_sort_rows([[r.get(c, "") for c in s1_cols] for r in teacher_rows], s1_cols)
    zb_write_sheet(wb_out, "教师基础数据", s1_cols, s1_data)

    s2_cols = ["姓名", "学科", "习题任务", "重练任务数", "重练覆盖率", "重练完成率", "重练得分率"]
    s2_data = zb_sort_rows([[r.get(c, "") for c in s2_cols] for r in teacher_rows], s2_cols)
    zb_write_sheet(wb_out, "错题重练数据", s2_cols, s2_data)

    s3_cols = ["姓名", "学科", "靶向任务", "靶向习题", "靶向完成率", "靶向班", "靶向学生", "打标签", "音频讲解", "视频讲解"]
    s3_data = zb_sort_rows([[r.get(c, "") for c in s3_cols] for r in teacher_rows], s3_cols)
    zb_write_sheet(wb_out, "靶向+资源建设", s3_cols, s3_data, auto_color=False)

    ws_c = wb_in["班级学习行为"]
    c_cols = zb_detect_cols(ws_c)
    s4_cols = ["班级", "习题任务", "师均习题任务", "生均题量", "完成率", "得分率",
               "阅卷率", "重练任务数", "重练覆盖率", "重练完成率", "重练得分率",
               "自学人次", "自学题量", "观看人次", "观看时长"]
    s4_data = []
    for r in range(4, ws_c.max_row + 1):
        subj = ws_c.cell(r, c_cols.get("学科", 3)).value
        if subj and str(subj).strip() == "总计":
            s4_data.append([ws_c.cell(r, c_cols.get(cn, 0)).value if cn in c_cols else "" for cn in s4_cols])
    zb_write_sheet(wb_out, "各班级数据", s4_cols, s4_data, auto_color=False)

    return wb_out


# ═══════════════════════════════════════════
# 功能2: 目标生考试知识点分析
# ═══════════════════════════════════════════

def mb_col_num_to_letter(col_num):
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + 65) + result
        col_num //= 26
    return result


def mb_normalize_exact(text):
    if text is None:
        return ""
    text = str(text)
    text = re.sub(r'^[\d\.\s\[\]]+', '', text)
    text = text.replace('（', '(').replace('）', ')')
    text = text.replace(' ', '').replace('的', '')
    return text.strip()


def mb_build_merged_cell_map(ws):
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        value = ws.cell(row=merged.min_row, column=merged.min_col).value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_map[(r, c)] = value
    return merged_map


def mb_step1_extract_score_rate(raw_bytes, target_class, max_q, fuzzy_enabled, fuzzy_threshold):
    file_buf = io.BytesIO(raw_bytes)

    def extract_tail_numbers(text):
        if pd.isna(text):
            return []
        text = str(text).strip()
        match = re.search(r'\s+(\d+)$', text)
        if not match:
            return []
        num_str = match.group(1)
        nums = []
        i = 0
        while i < len(num_str):
            if i + 2 <= len(num_str):
                two_digit = int(num_str[i:i + 2])
                if 10 <= two_digit <= max_q:
                    nums.append(two_digit)
                    i += 2
                    continue
            one_digit = int(num_str[i])
            if 1 <= one_digit <= 9:
                nums.append(one_digit)
            i += 1
        return nums

    df_raw = pd.read_excel(file_buf, sheet_name='知识点题号', header=None, usecols=[0])
    df_raw = df_raw.dropna()
    df_raw.columns = ['混合内容']

    mapping_records = []
    for _, row in df_raw.iterrows():
        mixed = row['混合内容']
        tihao_list = extract_tail_numbers(mixed)
        if tihao_list:
            knowledge = re.sub(r'\s+\d+$', '', mixed).strip()
            for th in tihao_list:
                mapping_records.append({'知识点': knowledge, '题号': th})

    df_mapping = pd.DataFrame(mapping_records)
    if len(df_mapping) == 0:
        return None, "知识点-题号映射为空！请检查知识点题号Sheet格式"

    file_buf.seek(0)
    df_qiangji = pd.read_excel(file_buf, sheet_name='强基临界生')
    all_classes = df_qiangji['班级'].dropna().unique()

    file_buf.seek(0)
    df_score = pd.read_excel(file_buf, sheet_name='小题分')
    df_score.columns = df_score.columns.astype(str)

    exclude_cols = {'序号', '姓名', '考号', '学校/年级', '班级'}
    tihao_cols = []
    for col in df_score.columns:
        if col in exclude_cols:
            continue
        try:
            num = int(float(col))
            if 1 <= num <= max_q:
                tihao_cols.append(col)
        except:
            pass

    tihao_cols = sorted(tihao_cols, key=lambda x: int(x))
    if len(tihao_cols) == 0:
        return None, "未识别到任何题号！请检查最大题号参数"

    for col in tihao_cols:
        df_score[col] = df_score[col].astype(str).str.rstrip('%').astype(float) / 100.0

    df_score['班级'] = df_score['班级'].astype(str)
    df_melt = df_score.melt(id_vars=['班级'], value_vars=tihao_cols, var_name='题号', value_name='得分率')
    df_melt['题号'] = df_melt['题号'].astype(int)
    df_avg = df_melt.groupby(['班级', '题号'], as_index=False)['得分率'].mean()

    existing_classes = df_avg['班级'].unique()
    target_classes = [str(cls) for cls in all_classes if str(cls) in existing_classes.astype(str)]

    output_buf = io.BytesIO()
    with pd.ExcelWriter(output_buf, engine='openpyxl') as writer:
        for cls in target_classes:
            cls_scores = df_avg[df_avg['班级'] == cls].set_index('题号')['得分率']
            records = []
            for _, row in df_mapping.iterrows():
                kp = row['知识点']
                q = row['题号']
                score = cls_scores.get(q)
                score_str = f"{score * 100:.1f}%" if pd.notna(score) else ''
                records.append({'知识点': kp, '题号': q, '得分率(%)': score_str})
            df_cls = pd.DataFrame(records)
            df_cls.to_excel(writer, sheet_name=cls, index=False)

    output_buf.seek(0)
    return output_buf, f"识别到 {len(target_classes)} 个班级"


def mb_step2_extract_from_graph(graph_bytes, max_scan_cols=40):
    wb = openpyxl.load_workbook(io.BytesIO(graph_bytes))

    ws = None
    for name in wb.sheetnames:
        if "班级图谱" in name:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    merged_map = mb_build_merged_cell_map(ws)

    def get_val(row, col):
        return merged_map.get((row, col), ws.cell(row=row, column=col).value)

    col_map = {}

    level_col = None
    for col in range(1, 10):
        val2 = str(get_val(2, col) or "").strip()
        if val2 == "层级":
            level_col = col
            col_map["层级"] = col
            break

    kp_start = None
    kp_end = None
    search_start = level_col + 1 if level_col else 2

    for col in range(search_start, 15):
        val2 = str(get_val(2, col) or "").strip()
        if val2 == "知识点":
            kp_start = col
            for col2 in range(col + 1, 20):
                val2_2 = str(get_val(2, col2) or "").strip()
                if val2_2 and val2_2 != "知识点":
                    kp_end = col2
                    break
            if kp_end is None:
                kp_end = col + 6
            break

    if kp_start is None:
        for col in range(search_start, 15):
            val1 = str(get_val(1, col) or "").strip()
            if "知识点" in val1:
                kp_start = col
                for col2 in range(col + 1, 20):
                    if str(get_val(1, col2) or "").strip():
                        kp_end = col2
                        break
                if kp_end is None:
                    kp_end = col + 6
                break

    if kp_start is None:
        return None, "未找到知识点列"

    col_map["知识点列范围"] = (kp_start, kp_end)

    for col in range(1, min(max_scan_cols, ws.max_column + 1)):
        val1 = str(get_val(1, col) or "").strip()
        val2 = str(get_val(2, col) or "").strip()

        if "21年" in val2 and "道" in val2: col_map["21年（道）"] = col
        elif "22年" in val2 and "道" in val2: col_map["22年（道）"] = col
        elif "23年" in val2 and "道" in val2: col_map["23年（道）"] = col
        elif "24年" in val2 and "道" in val2: col_map["24年（道）"] = col
        elif "25年" in val2 and "道" in val2: col_map["25年（道）"] = col
        elif "5年真题" in val2 and "道" in val2: col_map["5年真题（道）"] = col
        elif val2 == "作业": col_map["作业"] = col
        elif val2 == "测验": col_map["测验"] = col
        elif val2 == "考试": col_map["考试"] = col
        elif "图谱数据" in val1 and "年级" not in val1:
            for col2 in range(col, col + 6):
                if col2 > ws.max_column: break
                inner_val = str(get_val(2, col2) or "").strip()
                if inner_val == "得分率": col_map["得分率"] = col2
                elif inner_val == "题量": col_map["题量"] = col2

    all_data = []
    kp_map = {}

    for row in range(3, ws.max_row + 1):
        level = get_val(row, col_map.get("层级", 1))
        kp = None
        for col in range(kp_end - 1, kp_start - 1, -1):
            val = get_val(row, col)
            if val and isinstance(val, str) and len(val.strip()) > 0:
                kp = val
                break
        if kp is None:
            continue

        row_data = {
            "层级": level, "知识点": kp,
            "5年真题（道）": get_val(row, col_map.get("5年真题（道）", 9)),
            "21年（道）": get_val(row, col_map.get("21年（道）", 10)),
            "22年（道）": get_val(row, col_map.get("22年（道）", 11)),
            "23年（道）": get_val(row, col_map.get("23年（道）", 12)),
            "24年（道）": get_val(row, col_map.get("24年（道）", 13)),
            "25年（道）": get_val(row, col_map.get("25年（道）", 14)),
            "得分率（年级）": get_val(row, col_map.get("得分率", 15)),
            "题量（年级）": get_val(row, col_map.get("题量", 16)),
            "作业": get_val(row, col_map.get("作业", 17)),
            "测验": get_val(row, col_map.get("测验", 18)),
            "考试": get_val(row, col_map.get("考试", 19)),
            "行号": row,
        }
        all_data.append(row_data)
        norm_kp = mb_normalize_exact(str(kp))
        kp_map[norm_kp] = {'index': len(all_data) - 1, 'original': kp}

    return (all_data, kp_map, ws, col_map), "OK"


def mb_step3_match_score_rate(score_buf, all_data, kp_map, target_class, fuzzy_enabled, fuzzy_threshold):
    wb = openpyxl.load_workbook(score_buf)
    if target_class not in wb.sheetnames:
        return None, f"找不到班级Sheet: {target_class}，可用: {wb.sheetnames}"

    ws = wb[target_class]
    score_map = {}
    for row in range(2, ws.max_row + 1):
        kp = ws.cell(row=row, column=1).value
        score = ws.cell(row=row, column=3).value
        if kp and score:
            norm_kp = mb_normalize_exact(str(kp))
            score_map[norm_kp] = {'original': kp, 'score': str(score)}

    matched_exact = 0
    matched_fuzzy = 0

    for norm_kp, data in score_map.items():
        if norm_kp in kp_map:
            idx = kp_map[norm_kp]['index']
            all_data[idx]["本次考试得分率"] = data['score']
            matched_exact += 1

    if fuzzy_enabled:
        for norm_kp, data in score_map.items():
            if norm_kp in kp_map and "本次考试得分率" in all_data[kp_map[norm_kp]['index']]:
                continue
            best_sim = 0
            best_match = None
            for graph_kp, graph_data in kp_map.items():
                if "本次考试得分率" in all_data[graph_data['index']]:
                    continue
                sim = SequenceMatcher(None, norm_kp, graph_kp).ratio()
                if sim >= fuzzy_threshold and sim > best_sim:
                    best_sim = sim
                    best_match = graph_kp
            if best_match:
                idx = kp_map[best_match]['index']
                all_data[idx]["本次考试得分率"] = data['score']
                matched_fuzzy += 1

    total = len(score_map)
    return "本次考试得分率", f"精确匹配{matched_exact}个，模糊匹配{matched_fuzzy}个，未匹配{total - matched_exact - matched_fuzzy}个"


def mb_step4_generate_final(ws_graph, all_data, col_map):
    wb = ws_graph.parent
    for sheet_name in list(wb.sheetnames):
        if "班级图谱" not in sheet_name:
            del wb[sheet_name]

    ws2 = wb.create_sheet("整理好的得分率")
    header = ["层级", "知识点", "本次考试得分率", "得分率", "题量", "作业", "测验", "考试",
              "5年真题（道）", "21年（道）", "22年（道）", "23年（道）", "24年（道）", "25年（道）"]
    ws2.append(header)

    red_font = Font(color="FFFF0000")
    matched_data = [item for item in all_data if "本次考试得分率" in item]

    for item in matched_data:
        row_values = [
            item["层级"], item["知识点"], item["本次考试得分率"],
            item["得分率（年级）"], item["题量（年级）"], item["作业"],
            item["测验"], item["考试"], item["5年真题（道）"],
            item["21年（道）"], item["22年（道）"], item["23年（道）"],
            item["24年（道）"], item["25年（道）"],
        ]
        ws2.append(row_values)

        try:
            score_str = str(item["本次考试得分率"])
            if score_str.endswith('%'):
                score_val = float(score_str.rstrip('%'))
                if score_val < 80:
                    ws2.cell(row=ws2.max_row, column=2).font = red_font
                    ws2.cell(row=ws2.max_row, column=3).font = red_font
            grade_score_str = str(item["得分率（年级）"])
            if grade_score_str.endswith('%'):
                grade_score_val = float(grade_score_str.rstrip('%'))
                if grade_score_val < 80:
                    ws2.cell(row=ws2.max_row, column=4).font = red_font
        except:
            pass

    return wb


def mb_generate(raw_bytes, graph_bytes, target_class, max_q, fuzzy_enabled, fuzzy_threshold):
    score_buf, msg1 = mb_step1_extract_score_rate(raw_bytes, target_class, max_q, fuzzy_enabled, fuzzy_threshold)
    if score_buf is None:
        return None, msg1

    result, msg2 = mb_step2_extract_from_graph(graph_bytes)
    if result is None:
        return None, msg2

    all_data, kp_map, ws_graph, col_map = result

    score_col_name, msg3 = mb_step3_match_score_rate(score_buf, all_data, kp_map, target_class, fuzzy_enabled, fuzzy_threshold)
    if score_col_name is None:
        return None, msg3

    wb = mb_step4_generate_final(ws_graph, all_data, col_map)
    return wb, f"完成！匹配情况: {msg3}"


# ═══════════════════════════════════════════
# 功能3: 个人考试知识点分析
# ═══════════════════════════════════════════

def gr_parse_percentage(val):
    if pd.isna(val):
        return None
    if isinstance(val, str):
        val = val.replace('%', '').strip()
        if val in ['—', '-', '', '无']:
            return None
        try:
            return float(val)
        except:
            return None
    try:
        return float(val)
    except:
        return None


def gr_extract_topic_number(kp_name):
    if pd.isna(kp_name):
        return None, None
    kp_str = str(kp_name).strip().replace('\xa0', ' ')
    match = re.search(r'[\d]+(?:[、,]\d+)*$', kp_str)
    if match:
        topic_str = match.group().strip()
        clean_kp = kp_str[:match.start()].strip()
        if clean_kp:
            return topic_str, clean_kp
    return None, kp_str


def gr_extract_question_count(q_str):
    if pd.isna(q_str):
        return None
    match = re.search(r'共(\d+)道题', str(q_str))
    if match:
        return int(match.group(1))
    return None


def gr_smart_split_topic(topic_str):
    if topic_str is None:
        return None
    if '、' in topic_str or ',' in topic_str:
        return topic_str
    if topic_str.isdigit() and len(topic_str) > 2:
        if len(topic_str) == 4:
            a, b = topic_str[:2], topic_str[2:]
            if int(a) > 0 and int(b) > 0:
                return f"{a}、{b}"
        if len(topic_str) == 3:
            a, b = topic_str[0], topic_str[1:]
            if int(b) > 0:
                return f"{a}、{b}"
        if len(topic_str) == 2:
            a, b = topic_str[0], topic_str[1]
            if int(a) > 0 and int(b) > 0:
                return f"{a}、{b}"
    return topic_str


def gr_build_sheet2_index(df2):
    kp_cols = list(df2.columns[:7])
    cols = list(df2.columns)

    class_score_col = None
    class_qcount_col = None
    personal_score_col = None
    personal_qcount_col = None

    for i, col in enumerate(cols):
        col_str = str(col).strip()
        if class_score_col is None and (
                '图谱数据' in col_str or '年级' in col_str or '班级' in col_str) and '得分率' in str(df2.iloc[0, i]):
            class_score_col = col
            if i + 1 < len(cols):
                class_qcount_col = cols[i + 1]
        elif class_score_col is not None and personal_score_col is None:
            if ('图谱数据' in col_str or '个人' in col_str or '学生' in col_str) and '得分率' in str(df2.iloc[0, i]):
                personal_score_col = col
                if i + 1 < len(cols):
                    personal_qcount_col = cols[i + 1]

    if class_score_col is None or personal_score_col is None:
        for i, col in enumerate(cols):
            row0_val = str(df2.iloc[0, i]).strip() if pd.notna(df2.iloc[0, i]) else ''
            if row0_val == '得分率':
                if class_score_col is None:
                    class_score_col = col
                    if i + 1 < len(cols):
                        class_qcount_col = cols[i + 1]
                elif personal_score_col is None:
                    personal_score_col = col
                    if i + 1 < len(cols):
                        personal_qcount_col = cols[i + 1]

    data_start = 1 if len(df2) > 0 and str(df2.iloc[0, 0]).strip() == '层级' else 0

    sheet2_dict = {}
    for idx in range(data_start, len(df2)):
        row = df2.iloc[idx]
        for col in kp_cols:
            kp_name = row[col]
            if pd.notna(kp_name):
                kp_name_str = str(kp_name).strip()
                if kp_name_str and kp_name_str not in ['知识点', '层级', 'nan', '']:
                    if kp_name_str not in sheet2_dict:
                        class_rate = gr_parse_percentage(row.get(class_score_col)) if class_score_col else None
                        personal_rate = gr_parse_percentage(row.get(personal_score_col)) if personal_score_col else None
                        q_count = row.get(personal_qcount_col) if personal_qcount_col else None
                        if q_count is not None and pd.isna(q_count):
                            q_count = None

                        def normalize_rate(val):
                            if val is None: return None
                            return val if val > 1 else val * 100

                        sheet2_dict[kp_name_str] = {
                            'score_rate': normalize_rate(class_rate),
                            'personal_score_rate': normalize_rate(personal_rate),
                            'question_count': q_count
                        }

    return sheet2_dict


def gr_generate(input_bytes, output_sheet_name, only_with_topic):
    input_buf = io.BytesIO(input_bytes)
    df1 = pd.read_excel(input_buf, sheet_name='Sheet1')
    input_buf.seek(0)
    df2 = pd.read_excel(input_buf, sheet_name='Sheet2')

    sheet2_dict = gr_build_sheet2_index(df2)

    results = []
    for idx, row in df1.iterrows():
        kp_name = row.get('知识点', row.iloc[0])
        topic_str, clean_kp = gr_extract_topic_number(kp_name)
        if only_with_topic and topic_str is None:
            continue
        topic_num = gr_smart_split_topic(topic_str) if topic_str else None

        score_rate_col = '平均得分率' if '平均得分率' in df1.columns else df1.columns[1]
        score_rate = gr_parse_percentage(row.get(score_rate_col))
        exam_rate_str = f"{score_rate * 100:.1f}%" if score_rate is not None else "—"

        q_col = '涉及题目' if '涉及题目' in df1.columns else df1.columns[2]
        question_count = gr_extract_question_count(row.get(q_col))

        sheet2_data = sheet2_dict.get(clean_kp, {})
        class_rate = sheet2_data.get('score_rate')
        personal_rate = sheet2_data.get('personal_score_rate')
        sheet2_questions = sheet2_data.get('question_count')

        class_rate_str = f"{class_rate:.1f}%" if class_rate is not None else "—"
        personal_rate_str = f"{personal_rate:.1f}%" if personal_rate is not None else "—"
        q_count_val = float(sheet2_questions) if sheet2_questions is not None and not pd.isna(sheet2_questions) else "—"

        results.append({
            '题号': topic_num if topic_num else "—",
            '知识点': clean_kp,
            '考试正确率 (%)': exam_rate_str,
            '涉及题目 (道数)': float(question_count) if question_count is not None else "—",
            '班级日常训练得分率 (%)': class_rate_str,
            '个人得分率 (%)': personal_rate_str,
            '题量': q_count_val
        })

    df3 = pd.DataFrame(results)

    # 写入原文件
    input_buf.seek(0)
    wb = openpyxl.load_workbook(input_buf)

    existing_sheets = wb.sheetnames
    final_sheet_name = output_sheet_name
    if final_sheet_name in existing_sheets:
        n = 2
        while f"{final_sheet_name}_{n}" in existing_sheets:
            n += 1
        final_sheet_name = f"{final_sheet_name}_{n}"

    ws = wb.create_sheet(final_sheet_name)
    for ci, h in enumerate(df3.columns, 1):
        ws.cell(1, ci, h)
    for ri, row in enumerate(df3.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)

    return wb, f"生成 {len(results)} 行数据，Sheet名: {final_sheet_name}"


# ═══════════════════════════════════════════
# 功能4: 个人总表生成
# ═══════════════════════════════════════════

def zt_clean_percent_value(value):
    if pd.isna(value):
        return None
    value_str = str(value).strip()
    if '%' in value_str:
        value_str = value_str.replace('%', '')
    try:
        return float(value_str)
    except:
        return None


def zt_load_student_data(sheet_name, excel_file):
    try:
        df = excel_file.parse(sheet_name)
        cleaned_df = pd.DataFrame()
        cleaned_df['题号'] = df['题号'].astype(str)
        cleaned_df['知识点'] = df['知识点']

        for col in df.columns:
            col_lower = str(col).lower()
            if ('考试' in col_lower or 'exam' in col_lower) and ('正确率' in col_lower or 'correct' in col_lower):
                cleaned_df[f'{sheet_name}考试正确率'] = df[col].apply(zt_clean_percent_value)
            elif ('个人' in col_lower or 'personal' in col_lower) and ('得分率' in col_lower or 'score' in col_lower):
                cleaned_df[f'{sheet_name}个人得分率'] = df[col].apply(zt_clean_percent_value)
            elif ('班级' in col_lower or 'class' in col_lower) and ('得分率' in col_lower or 'correct' in col_lower):
                cleaned_df['班级日常训练得分率'] = df[col].apply(zt_clean_percent_value)

        if f'{sheet_name}考试正确率' not in cleaned_df.columns:
            for col in df.columns:
                if '正确率' in str(col):
                    cleaned_df[f'{sheet_name}考试正确率'] = df[col].apply(zt_clean_percent_value)
                    break
        if f'{sheet_name}个人得分率' not in cleaned_df.columns:
            for col in df.columns:
                if '得分率' in str(col):
                    cleaned_df[f'{sheet_name}个人得分率'] = df[col].apply(zt_clean_percent_value)
                    break

        return cleaned_df
    except Exception as e:
        return pd.DataFrame()


def zt_generate(input_bytes, subject_name):
    input_buf = io.BytesIO(input_bytes)
    excel_file = pd.ExcelFile(input_buf)

    all_sheets = excel_file.sheet_names
    student_sheets = [s for s in all_sheets if s != '总表']

    if not student_sheets:
        return None, "未找到学生表"

    student_dfs = {}
    valid_students = []
    for sheet_name in student_sheets:
        df = zt_load_student_data(sheet_name, excel_file)
        if not df.empty and '题号' in df.columns and len(df) > 0:
            student_dfs[sheet_name] = df
            valid_students.append(sheet_name)

    if not valid_students:
        return None, "没有有效学生表"

    all_questions = {}
    if '总表' in all_sheets:
        try:
            original_df = excel_file.parse('总表', header=None)
            for i in range(3, original_df.shape[0]):
                row_data = original_df.iloc[i]
                if pd.isna(row_data[0]) or str(row_data[0]).strip() == '':
                    continue
                q_num = str(row_data[0]).strip()
                lesson_name = row_data[1]
                if pd.isna(lesson_name) or str(lesson_name).strip() == '':
                    continue
                lesson_name = str(lesson_name).strip()
                all_questions[(q_num, lesson_name)] = {
                    '题号': q_num, '知识点': lesson_name, '班级日常训练得分率': None
                }
        except:
            pass

    if not all_questions:
        for sheet_name in valid_students:
            student_df = student_dfs[sheet_name]
            for _, row in student_df.iterrows():
                if pd.isna(row['题号']) or str(row['题号']).strip() == '':
                    continue
                q_num = str(row['题号']).strip()
                lesson_name = str(row['知识点']).strip() if pd.notna(row['知识点']) else '未命名知识点'
                key = (q_num, lesson_name)
                if key not in all_questions:
                    all_questions[key] = {
                        '题号': q_num, '知识点': lesson_name,
                        '班级日常训练得分率': row.get('班级日常训练得分率')
                    }

    summary_data = []
    for key in all_questions:
        record = all_questions[key].copy()
        for sheet_name in valid_students:
            record[f'{sheet_name}考试正确率'] = None
            record[f'{sheet_name}个人得分率'] = None

        for sheet_name in valid_students:
            student_df = student_dfs[sheet_name]
            matching = student_df[
                (student_df['题号'] == key[0]) &
                (student_df['知识点'].str.contains(key[1], na=False))
            ]
            if not matching.empty:
                match = matching.iloc[0]
                if record['班级日常训练得分率'] is None and '班级日常训练得分率' in match:
                    record['班级日常训练得分率'] = match['班级日常训练得分率']
                if f'{sheet_name}考试正确率' in match:
                    record[f'{sheet_name}考试正确率'] = match[f'{sheet_name}考试正确率']
                if f'{sheet_name}个人得分率' in match:
                    record[f'{sheet_name}个人得分率'] = match[f'{sheet_name}个人得分率']
            else:
                matching = student_df[student_df['题号'] == key[0]]
                if len(matching) == 1:
                    match = matching.iloc[0]
                    if record['班级日常训练得分率'] is None and '班级日常训练得分率' in match:
                        record['班级日常训练得分率'] = match['班级日常训练得分率']
                    if f'{sheet_name}考试正确率' in match:
                        record[f'{sheet_name}考试正确率'] = match[f'{sheet_name}考试正确率']
                    if f'{sheet_name}个人得分率' in match:
                        record[f'{sheet_name}个人得分率'] = match[f'{sheet_name}个人得分率']

        summary_data.append(record)

    summary_df = pd.DataFrame(summary_data)

    def natural_sort_key(s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split('(\\d+)', s)]

    try:
        if not summary_df.empty and '题号' in summary_df.columns:
            summary_df['sort_key'] = summary_df['题号'].apply(natural_sort_key)
            summary_df = summary_df.sort_values('sort_key').drop('sort_key', axis=1).reset_index(drop=True)
    except:
        pass

    # 生成带格式的Excel
    input_buf.seek(0)
    wb = openpyxl.load_workbook(input_buf)
    if '总表' in wb.sheetnames:
        del wb['总表']

    ws = wb.create_sheet('总表', 0)
    title = f'目标生{subject_name}考试知识点对比分析情况'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + 2 * len(valid_students))
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = '题号'
    ws['B2'] = '知识点'
    ws['C2'] = '班级日常训练得分率'

    col_idx = 4
    for sheet_name in valid_students:
        ws.cell(row=2, column=col_idx, value=sheet_name)
        col_idx += 2

    col_idx = 4
    for sheet_name in valid_students:
        ws.cell(row=3, column=col_idx, value='考试正确率')
        ws.cell(row=3, column=col_idx + 1, value='个人得分率')
        col_idx += 2

    header_font = Font(size=11, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=3 + 2 * len(valid_students)):
        for cell in row:
            if cell.value is not None:
                cell.font = header_font
                cell.alignment = center_alignment

    if summary_df is not None and len(summary_df) > 0:
        for idx, row in summary_df.iterrows():
            row_num = idx + 4
            ws.cell(row=row_num, column=1, value=row['题号'])
            ws.cell(row=row_num, column=2, value=row['知识点'])
            class_score = row['班级日常训练得分率']
            if pd.notna(class_score):
                if class_score > 1:
                    ws.cell(row=row_num, column=3, value=float(class_score) / 100)
                else:
                    ws.cell(row=row_num, column=3, value=float(class_score))

            col_idx = 4
            for sheet_name in valid_students:
                exam_col = f'{sheet_name}考试正确率'
                if exam_col in row and pd.notna(row[exam_col]):
                    ws.cell(row=row_num, column=col_idx, value=float(row[exam_col]))
                score_col = f'{sheet_name}个人得分率'
                if score_col in row and pd.notna(row[score_col]):
                    ws.cell(row=row_num, column=col_idx + 1, value=float(row[score_col]))
                col_idx += 2

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    max_row = max(4, ws.max_row)
    for row in ws.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=3 + 2 * len(valid_students)):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
            if cell.column == 3 and cell.value is not None:
                cell.number_format = '0.0%'
            elif cell.column > 3 and cell.value is not None:
                try:
                    float(cell.value)
                    cell.number_format = '0.0'
                except:
                    pass

    return wb, f"包含 {len(valid_students)} 个学生，{len(summary_df)} 个知识点"


# ═══════════════════════════════════════════
# 页面1: 周报数据生成
# ═══════════════════════════════════════════

if page == "📊 周报数据生成":
    st.title("📊 周报数据生成")
    st.markdown("上传教学报告Excel，一键生成带条件格式的教师数据统计表")

    uploaded = st.file_uploader("📎 上传教学报告Excel文件", type=['xlsx', 'xls'], key="zb")
    if uploaded:
        with st.spinner("正在处理..."):
            try:
                wb = zb_generate(uploaded.read())
                excel_bytes = wb_to_bytes(wb)
                st.success("✅ 处理完成！")
                st.download_button("📥 下载统计表Excel", data=excel_bytes,
                                   file_name="教师数据统计表_生成.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ 处理失败：{e}")
                st.info("请确认上传的是包含「教师教学详情」和「班级学习行为」Sheet的教学报告文件")


# ═══════════════════════════════════════════
# 页面2: 目标生考试知识点分析
# ═══════════════════════════════════════════

elif page == "🎯 目标生考试知识点分析":
    st.title("🎯 目标生考试知识点分析")
    st.markdown("上传小题分文件和班级图谱，生成目标生考试知识点分析表")

    raw_file = st.file_uploader("📎 上传小题分平均得分率提取文件", type=['xlsx', 'xls'], key="mb1")
    graph_file = st.file_uploader("📎 上传班级图谱文件", type=['xlsx', 'xls'], key="mb2")

    col1, col2 = st.columns(2)
    with col1:
        target_class = st.text_input("目标班级名称", value="清北2班")
    with col2:
        max_q = st.number_input("当前最大题号", min_value=1, max_value=50, value=20)

    col3, col4 = st.columns(2)
    with col3:
        fuzzy_enabled = st.checkbox("开启模糊匹配", value=True)
    with col4:
        fuzzy_threshold = st.slider("模糊匹配阈值", 0.5, 1.0, 0.85, 0.05) if fuzzy_enabled else 0.85

    if raw_file and graph_file and st.button("🚀 开始生成"):
        with st.spinner("正在处理..."):
            try:
                wb, msg = mb_generate(raw_file.read(), graph_file.read(),
                                      target_class, max_q, fuzzy_enabled, fuzzy_threshold)
                if wb is None:
                    st.error(f"❌ {msg}")
                else:
                    excel_bytes = wb_to_bytes(wb)
                    st.success(f"✅ {msg}")
                    st.download_button("📥 下载分析结果Excel", data=excel_bytes,
                                       file_name=f"{target_class}_考试知识点分析.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ 处理失败：{e}")


# ═══════════════════════════════════════════
# 页面3: 个人考试知识点分析
# ═══════════════════════════════════════════

elif page == "📝 个人考试知识点分析":
    st.title("📝 个人考试知识点分析")
    st.markdown("上传个人知识点图谱文件（含Sheet1和Sheet2），生成个人考试知识点分析")

    gr_file = st.file_uploader("📎 上传个人知识点Excel文件", type=['xlsx', 'xls'], key="gr")
    col1, col2 = st.columns(2)
    with col1:
        output_sheet_name = st.text_input("输出Sheet名称", value="分析结果")
    with col2:
        only_with_topic = st.checkbox("只保留有题号的行", value=True)

    if gr_file and st.button("🚀 开始生成"):
        with st.spinner("正在处理..."):
            try:
                wb, msg = gr_generate(gr_file.read(), output_sheet_name, only_with_topic)
                if wb is None:
                    st.error(f"❌ {msg}")
                else:
                    excel_bytes = wb_to_bytes(wb)
                    st.success(f"✅ {msg}")
                    st.download_button("📥 下载分析结果Excel", data=excel_bytes,
                                       file_name="个人考试知识点分析.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ 处理失败：{e}")


# ═══════════════════════════════════════════
# 页面4: 个人总表生成
# ═══════════════════════════════════════════

elif page == "📈 个人总表生成":
    st.title("📈 个人总表生成")
    st.markdown("上传含多个学生Sheet的Excel文件，生成对比分析总表")

    zt_file = st.file_uploader("📎 上传学生数据Excel文件", type=['xlsx', 'xls'], key="zt")
    subject_name = st.text_input("科目名称", value="历史")

    if zt_file and st.button("🚀 开始生成"):
        with st.spinner("正在处理..."):
            try:
                wb, msg = zt_generate(zt_file.read(), subject_name)
                if wb is None:
                    st.error(f"❌ {msg}")
                else:
                    excel_bytes = wb_to_bytes(wb)
                    st.success(f"✅ {msg}")
                    st.download_button("📥 下载总表Excel", data=excel_bytes,
                                       file_name=f"{subject_name}_考试分析总表.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ 处理失败：{e}")
